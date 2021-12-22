
import time
import os
import shutil
import skimage
import openpyxl
import torch
import torch.nn as nn
from fastreid.data.build import (build_reid_test_data_loader,
                                 build_reid_query_data_loader,
                                 build_reid_att_query_data_loader,
                                 build_reid_gallery_data_loader,
                                 build_reid_train_loader,)
from fastreid.engine import DefaultTrainer
from fastreid.solver.build import build_optimizer
from fastreid.utils.checkpoint import Checkpointer
from openpyxl.utils import get_column_letter
from torchvision import transforms
from fastreid.engine import DefaultTrainer

device= 'cuda'
excel_name = 'result.xlsx'

Attack_algorithm_library=['FGSM','IFGSM','MIFGSM','ODFA','MISR','FNA','MUAP','SSAE']
Defense_algorithm_library=['ADV','GOAT','EST','SES','PNP']

evaluation_indicator=['Rank-1','mAP',"TPR@FPR={:.0e}".format(1e-2)]
evaluation_attackIndex= ['DmAP','SDSIM','AttackIndex']
evaluation_defenseIndex =['def-SDSIM','DefenseIndex']
attack_R_type = ['QA+','QA-','GA+','GA-']

def get_query_set(cfg,relabel=True):
    query_set=build_reid_query_data_loader(cfg,cfg.DATASETS.TESTS[0],relabel=relabel)
    return query_set

def get_att_query_set(cfg,relabel=True):
    adv_query_set = build_reid_att_query_data_loader(cfg,cfg.DATASETS.TESTS[0],relabel=relabel)
    return adv_query_set

def get_test_set(cfg,relabel=True):
    test_set = build_reid_test_data_loader(cfg,cfg.DATASETS.TESTS[0],relabel=relabel)
    return test_set

def get_gallery_set(cfg,relabel=True):
    gallery_set=build_reid_gallery_data_loader(cfg,cfg.DATASETS.TESTS[0],relabel=relabel)
    return gallery_set

def get_train_set(cfg):
    train_data_loader = build_reid_train_loader(cfg)
    return train_data_loader

@torch.no_grad()
def eval_train(model,data_loader,max_id=-1):
    print('*****************evluation*****************')
    model.eval()
    correct = 0 
    softmax = nn.Softmax(dim=1)
    n=0
    for id,data in enumerate(data_loader):
        if max_id!=-1 and id>max_id:
            break
        logits = model(data)
        probabilities = softmax(logits)
        pred = probabilities.argmax(dim=1,keepdim=True)
        targets = data['targets'].to(device)
        correct += pred.eq(targets.view_as(pred)).sum().item()
        n+=targets.shape[0]
    return 100. * correct / n

def eval_ssim(images1,images2):
    size = images1.shape[0]
    SSIM = 0
    for i in range(size):
        image1 = CHW_to_HWC(images1[i])
        image2 = CHW_to_HWC(images2[i])
        SSIM += skimage.measure.compare_ssim(image1,image2,multichannel=True)
    SSIM/=size
    return SSIM

def save_image(imgs,paths,str):
      
    imgs= imgs.cpu()                                                              
    toPIL = transforms.ToPILImage() #这个函数可以将张量转为PIL图片，由小数转为0-255之间的像素值
    for img,path in zip(imgs,paths):
        pic = toPIL(img)
        position = path.find('query/') # path给的是绝对位置
        name = path[position+6:] #只需要提取出其名字即可
        pic.save(path[:position]+str+'/'+name)

    print(f"successful save in {path[:position]+str}")

def classify_test_set(cfg,data_loader):
    '''
    固定住backbone的权重，利用query_set重新训练cls_layer的weight,使其也能准确的分类图像和id
    转换成分类问题，可较好的接收针对于分类问题的算法攻击
    
    '''
    #set the number of neurons of the classifier layer as query set's targets number
    cfg = DefaultTrainer.auto_scale_hyperparams(cfg,data_loader.dataset.num_classes)

    model = DefaultTrainer.build_model_main(cfg)  # use baseline_train
    Checkpointer(model).load(cfg.MODEL.WEIGHTS)  # load trained model

    optimizer = build_optimizer(cfg,model) 
    loss_fun = nn.CrossEntropyLoss()
    #fix the weight of the backbone layers,and we only train the classifier for query set
    for _,parm in enumerate(model.parameters()):
        parm.requires_grad=False
    for _,parm in enumerate(model.heads.classifier.parameters()):
        parm.requires_grad=True

    # the origin model was trained by training set ,which has different targets number from query set,
    # so the number of neurons of the classifier layer was different, the weight couldn'y directly be loaded,
    # and you can see the warning like 
    # "Skip loading parameter 'heads.classifier.weight' to the model due to incompatible shapes: (751, 2048) in the checkpoint but (0, 2048) in the model! You might want to double check if this is expected.'
    # so we skip this part of weight and initial weights as below
    torch.nn.init.xavier_normal_(model.heads.classifier.weight)

    epoch = 5  #as far as i noticed,within 5 epoches the classifier layer can be trained to efficiently behave well
    for i in range(epoch):
        model.train()
        for _,data in enumerate(data_loader):
            targets = data['targets'].to(device)
            optimizer.zero_grad()
            logits = model(data)
            loss = loss_fun(logits,targets)
            loss.backward()
            optimizer.step()
            
        accurency = eval_train(model,data_loader) #show the accurrency
        print('The accurency of query set in Train Epoch {} is {}%'.format(i,accurency))
        print('---------------------------------------------------------------------')
    Checkpointer(model,'model').save('test_trained') 
    # model was saved in ./model/test_trained.pth
    # cfg.TESTSET_TRAINED_WEIGHT
    print('the model was saved in ./model/test_trained.pth')

def change_preprocess_image(cfg):
    def preprocess_image(batched_inputs):
        """
        Normalize and batch the input images.
        """
        if isinstance(batched_inputs, dict):
            images = batched_inputs["images"].to(device)
        elif isinstance(batched_inputs, torch.Tensor):
            images = batched_inputs.to(device)
        else:
            raise TypeError("batched_inputs must be dict or torch.Tensor, but get {}".format(type(batched_inputs)))
        
        images = torch.mul(images,255)
        images = torch.sub(images,torch.tensor(cfg.MODEL.PIXEL_MEAN).view(1, -1, 1, 1).to(device))
        images = torch.div(images,torch.tensor(cfg.MODEL.PIXEL_STD).view(1, -1, 1, 1).to(device))
        return images
    return preprocess_image
    
def get_result(cfg,model_path,step:str)->dict:
    model =DefaultTrainer.build_model(cfg)  # 启用baseline,用于测评
    Checkpointer(model).load(model_path)
    if step=='attack':
        result = DefaultTrainer.advtest(cfg,model)# advtest用于测试adv_query与gallery合成的test_set
    elif step=='pure' or step=='defense':
        result = DefaultTrainer.test(cfg,model)# test用于测试query与gallery合成的test_set

    return result

def check(save_pos):
    return os.path.exists(save_pos)

def remove_Sheet(wb,sheet_list):
    if 'Sheet' in sheet_list:
        wb.remove(wb['Sheet'])
    if 'Sheet1' in sheet_list:
        wb.remove(wb['Sheet1'])

def sheet_init(sheet):
    for row in range (3,4+len(Attack_algorithm_library)*len(evaluation_indicator)*len(attack_R_type)):
        for col in range (2,7+len(Defense_algorithm_library)*2):
            sheet.column_dimensions[get_column_letter(col)].width = 20
            sheet.row_dimensions[row].height = 40
    sheet['E3']='PURE VALUE'
    sheet['F3']='AFTER ATTACK'
    for row,name in enumerate(Attack_algorithm_library):
        row = len(evaluation_indicator)*len(attack_R_type)*row
        sheet['B'+str(4+row)]=name
        for bias,type in enumerate(attack_R_type):
            bias = len(evaluation_indicator)*bias
            sheet['C'+str(4+row+bias)]=type
            for i,indicator in enumerate(evaluation_indicator):
                sheet['D'+str(4+row+bias+i)]=indicator
            

    for col,name in enumerate(Defense_algorithm_library):
        sheet[get_column_letter(7+2*col)+str(3)]=name
        sheet[get_column_letter(8+2*col)+str(3)]='ATFER ATTACK'

def sheet_index_init(sheet_index):
    for row in range (3,4+4*len(Attack_algorithm_library)):
        for col in range (2,6+2*len(Defense_algorithm_library)):
            sheet_index.column_dimensions[get_column_letter(col)].width = 20
            sheet_index.row_dimensions[row].height = 40

    bias = 4
    for row,name in enumerate(Attack_algorithm_library):
        for i in range(4):
            sheet_index['B'+str(bias+i+4*row)]=name
        for i,type in enumerate(attack_R_type):
            sheet_index['C'+str(bias+4*row+i)] =type
    
    for i in range(len(evaluation_attackIndex)):
        sheet_index[get_column_letter(i+4)+'3'] = evaluation_attackIndex[i]

    for i in range(len(Defense_algorithm_library)):
        line = 4+len(evaluation_attackIndex)+len(evaluation_defenseIndex)*i
        for j in range(len(evaluation_defenseIndex)-1):
            sheet_index[get_column_letter(j+line)+'3'] = evaluation_defenseIndex[j]
        sheet_index[get_column_letter(len(evaluation_defenseIndex)-1+line)+'3'] = Defense_algorithm_library[i]
        

def save_data(cfg,pure_result,att_result,def_result,def_adv_result,sheet):

    row = 4 + len(evaluation_indicator)*len(attack_R_type)*Attack_algorithm_library.index(cfg.ATTACKMETHOD)
    row += len(evaluation_indicator)*attack_R_type.index(cfg.ATTACKTYPE+cfg.ATTACKDIRECTION)
    column1 = chr(ord('G') + Defense_algorithm_library.index(cfg.DEFENSEMETHOD)*2)
    column2 = chr(ord(column1)+1)

    for col,dict_name in {'E':pure_result,'F':att_result,column1:def_result,column2:def_adv_result}.items():
        for i in range(len(evaluation_indicator)):
            sheet[col+str(row+i)] = round(dict_name[evaluation_indicator[i]],2)

def CalculateIndex(cfg,pure_result,att_result,def_result,def_adv_result,SSIM,def_SSIM,sheet):

    row_start = 4 + 4*Attack_algorithm_library.index(cfg.ATTACKMETHOD)
    bias = attack_R_type.index(cfg.ATTACKTYPE+cfg.ATTACKDIRECTION)
    row = row_start+bias
    
    AttackIndex = {}
    DefenseIndex = {}
    # AttackIndex calculate
    AttackIndex['DmAP'] = pure_result['mAP']-att_result['mAP']
    AttackIndex['SDSIM'] = 1-SSIM
    AttackIndex['AttackIndex'] = AttackIndex['DmAP']/AttackIndex['SDSIM']
    
    # DefenseIndex calculate
    DefenseIndex['def-SDSIM'] = 1-def_SSIM
    Delta_mAP_keep = pure_result['mAP']-def_result['mAP']
    Delta_mAP_improve = def_adv_result['mAP']-att_result['mAP']
    Delta_SSIM = def_SSIM - SSIM
    DefenseIndex['DefenseIndex'] = Delta_mAP_improve*Delta_SSIM/Delta_mAP_keep

    for i,indicator in enumerate(evaluation_attackIndex):
        sheet[get_column_letter(4+i)+str(row)] = round(AttackIndex[indicator],2)
    
    line = 4+len(evaluation_attackIndex)+len(evaluation_defenseIndex)*Defense_algorithm_library.index(cfg.DEFENSEMETHOD)
    for i in range(len(evaluation_defenseIndex)-1):
        sheet[get_column_letter(line+i)+str(row)] = DefenseIndex[evaluation_defenseIndex[i]]
    sheet[get_column_letter(line+len(evaluation_defenseIndex)-1)+str(row)] = round(DefenseIndex['DefenseIndex'],2)


def record(cfg,pure_result,att_result,def_result,def_adv_result,SSIM,def_SSIM,save_pos= excel_name):
    if check(save_pos):
        wb = openpyxl.load_workbook(save_pos)
    else :
        wb = openpyxl.Workbook()
    sheet_list = wb.sheetnames
    sheet_name = f'{cfg.DATASETS.NAMES[0]}_{cfg.CFGTYPE}'
    if sheet_name in sheet_list:
        sheet = wb[sheet_name]
    else :
        sheet = wb.create_sheet(title = sheet_name)
    
    sheet_attackIndex_name=f'Index_{cfg.DATASETS.NAMES[0]}_{cfg.CFGTYPE}'
    if sheet_attackIndex_name in sheet_list:
        sheet_index = wb[sheet_attackIndex_name]
    else :
        sheet_index = wb.create_sheet(title = sheet_attackIndex_name)
    sheet_init(sheet)
    sheet_index_init(sheet_index)
    if def_result != None and def_adv_result != None and pure_result!=None and att_result!=None:
        save_data(cfg,pure_result,att_result,def_result,def_adv_result,sheet)
        CalculateIndex(cfg,pure_result,att_result,def_result,def_adv_result,SSIM,def_SSIM,sheet_index)
    #save_config(cfg,pure_result,adv_result,def_adv_result,sheet)
    remove_Sheet(wb,sheet_list)
    wb.save(save_pos)

def process_set(loader, model,device='cuda'):
    """Extract all the features of images from a loader using a model.
    
    Arguments:
        loader {Pytorch dataloader} -- loader of the images
        model {Pytorch model} -- model used to extract the features
        device {cuda device} -- 
    
    Returns:
        features -- Tensor of the features of the queries
        ids -- numpy array of ids
        cams -- numpy array of camera ids
    """

    ids = []
    cams = []
    features = []
    model.eval()
    for id, data in enumerate(loader):
        with torch.no_grad():
            output = model((data['images']/255.0).to(device))
        features.append(output)
        ids.append(data['targets'].cpu())
        cams.append(data['camids'].cpu())
    ids = torch.cat(ids, 0)
    cams = torch.cat(cams, 0)
    features = torch.cat(features, 0)
    return features.cpu(), ids.numpy(), cams.numpy()

def pairwise_distance(x, y):
    """Compute the matrix of pairwise distances between tensors x and y

    Args:
        x (Tensor)
        y (Tensor)

    Returns:
        Tensor: matrix of pairwise distances
    """
    m, n = x.size(0), y.size(0)
    x = x.view(m, -1)
    y = y.view(n, -1)
    dist = torch.pow(x, 2).sum(dim=1, keepdim=True).expand(m, n) + \
        torch.pow(y, 2).sum(dim=1, keepdim=True).expand(n, m).t()
    dist.addmm_(x, y.t(), beta=1, alpha=-2)
    dist = dist.clamp(min=1e-12).sqrt()
    return dist

def CHW_to_HWC(images):
    # compare_ssim function in sklearn is used for HWC images
    # this function was used to transform CHW images to HWC images
    # (size)3x256x128 -> (size)256x128x3
    images = images.cpu()
    images = images.permute(1,2,0)
    # tensor.permute()
    # change dimension of tensor, only tensor
    # previous dimension C(0),H(1),W(2) to H(1),W(2),C(0) 
    return images.detach().numpy()

def mkdir(path):
    folder = os.path.exists(path)
    if not folder:
        print("dirtory "+path+" has been created")
    else:
        if os.path.exists(path+'/origin'):
            shutil.rmtree(path+"/origin")
        if os.path.exists(path+'/attack'):
            shutil.rmtree(path+"/attack")
        if os.path.exists(path+'/defense'):
            shutil.rmtree(path+"/defense")
        if os.path.exists(path+'/adv_defense'):
            shutil.rmtree(path+"/adv_defense")
    os.makedirs(path+"/origin") 
    os.makedirs(path+"/attack") 
    os.makedirs(path+"/defense")
    os.makedirs(path+"/adv_defense")

def record_order(cfg,pure_result,att_result,def_result,def_adv_result, pictureNumber = 50,save_pic=False):

    batch_size = cfg.TEST.IMS_PER_BATCH

    # query_set = get_query_set(cfg,relabel=False)
    # gallery_set = get_gallery_set(cfg,relabel=False)
    
    # gallery_images = []
    # gallery_targets = []
    # for _ ,data in enumerate(gallery_set):
    #     gallery_images.append(data['images'].cpu()/255.0)
    #     gallery_targets.append(data['targets'])

    test_dataset,num_query = DefaultTrainer.build_test_loader(cfg,dataset_name=cfg.DATASETS.NAMES[0])
    images = []
    pids = []
    camids = []
    print('batch size = ',batch_size)
    print('num_query = ',num_query)

    for _,data in enumerate(test_dataset):
        images.append((data['images']/255.0).cpu())
        pids.append(data['targets'].cpu())
        camids.append(data['camids'].cpu())
    
    images = torch.cat(images,dim=0)
    pids = torch.cat(pids, dim=0)
    camids = torch.cat(camids,dim=0)

    query_images = images[:num_query]
    query_pids = pids[:num_query]

    gallery_images = images[num_query:]
    gallery_pids = pids[num_query:]

    result_order = {}
    q_pid_save = {}
    g_pids_save = {}
    for name,dict in {'origin':pure_result,'attack':att_result,'defense':def_result,'adv_defense':def_adv_result}.items():
        result_order[name]=[]
        q_pid_save[name]=None
        g_pids_save[name]=[]
        if dict!=None:
            for key,list in {'result_order':result_order,'q_pid_save':q_pid_save,'g_pids_save':g_pids_save}.items():
                list[name]=dict[key]

    pictureNumber = pictureNumber  
    #the same as 'max_rank' in engine/evaluation/rank.py  -> function(eval_cuhk/eval_market1501)

    
    path = f"{os.getcwd()}/logs_pic/{cfg.DATASETS.NAMES[0]}/{cfg.ATTACKTYPE}{cfg.ATTACKDIRECTION}/{cfg.ATTACKMETHOD}_{cfg.DEFENSEMETHOD}"
    mkdir(path)

    log(cfg,q_pid_save,g_pids_save,pictureNumber)
    
    
    # start to save pictures

    # save method 1
    # toPIL = transforms.ToPILImage()

    # for name,pos in result_order.items():
    #     print(f'start to save pictures of query&gallery set of {name}')
    #     for _,data in enumerate(query_set):
    #         for i in range(pictureNumber):
    #             img = (data['images'].cpu()/255.0)[i]
    #             target = data['targets'][i].item()
    #             os.makedirs(f'{path}/{name}/{i}')
    #             toPIL(img).save(f'{path}/{name}/{i}/q_{target}.jpg')
    #             if q_pid_save[name]==None:
    #                 continue
    #             if target!=q_pid_save[name][i]:
    #                 print("target =",target)
    #                 print('q_pid_save = ',q_pid_save[name][i])
    #         break
    #     print('query set pictures are all saved ')
    #     if pos==[]:
    #         continue
    #     for i in range(pictureNumber):
    #         for j in range(pictureNumber):
    #             toPIL(gallery_images[pos[i][j]//batch_size][pos[i][j]%batch_size]).save(f'{path}/{name}/{i}/{j}_{gallery_targets[pos[i][j]//batch_size][pos[i][j]%batch_size].item()}.jpg')
    #             if gallery_targets[pos[i][j]//batch_size][pos[i][j]%batch_size].item()!=g_pids_save[name][i][j]:
    #                 print("target =",gallery_targets[pos[i][j]//batch_size][pos[i][j]%batch_size].item())
    #                 print('g_pid_save = ',g_pids_save[name][i][j])

    #     print('gallery set pictures are all saved')



    # save method 2
    if save_pic:
        toPIL = transforms.ToPILImage()
        for name,pos in result_order.items():
            print(f'start to save pictures of query&gallery set of {name}')
        
            for i in range(pictureNumber):
                img = query_images[i]
                target = query_pids[i]
                os.makedirs(f'{path}/{name}/{i}')
                toPIL(img).save(f'{path}/{name}/{i}/q_{target}.jpg')
                if q_pid_save[name]==None:
                    continue
                if target!=q_pid_save[name][i]:
                    print("target =",target)
                    print('q_pid_save = ',q_pid_save[name][i])
        
            print('query set pictures are all saved ')
            if pos==[]:
                continue
            for i in range(pictureNumber):
                for j in range(pictureNumber):
                    toPIL(gallery_images[pos[i][j]]).save(f'{path}/{name}/{i}/{j}_{gallery_pids[pos[i][j]].item()}.jpg')
                    if gallery_pids[pos[i][j]].item()!=g_pids_save[name][i][j]:
                        print("target =",gallery_pids[pos[i][j]].item())
                        print('g_pid_save = ',g_pids_save[name][i][j])

            print('gallery set pictures are all saved')

    else :
        print('You choose not to save pictures')

def log(cfg):
    path = f"{os.getcwd()}/logs_pic"

    file = open(path+"/log.txt",'a')# write from the end of the txt,so it will record all your jobs
    file.write('---------------------------------start-log---------------------------------\n')
    time_info = time.asctime( time.localtime(time.time()))
    file.write("the log time is "+time_info+"\n\n") 

    file.write('Configuration:\n')
    file.write(f"     dataset: {cfg.DATASETS.NAMES[0]}\n")
    file.write(f"     attack : {cfg.ATTACKMETHOD!=None}\n")
    if cfg.ATTACKMETHOD!=None:
        file.write(f"            attack  method    = {cfg.ATTACKMETHOD}\n")
        file.write(f"            attack  type      = {cfg.ATTACKTYPE}\n")
        file.write(f"            attack  direction = {cfg.ATTACKDIRECTION}\n")
        file.write(f"            attack  direction = {cfg.ATTACKDIRECTION}\n")
    file.write(f"     defense: {cfg.DEFENSEMETHOD!=None} \n")
    if cfg.DEFENSEMETHOD!=None:
        file.write(f"            defense method    = {cfg.DEFENSEMETHOD}\n"  )
    
    file.write('\n---------------------------------end-log---------------------------------\n')
    file.write('|||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||\n')
    file.close()
    print("log over")

def print_info(str):
    print('---------------------------------------------------------')
    print('---------------------------------------------------------')
    print(str)
    print('---------------------------------------------------------')
    print('---------------------------------------------------------')

def analyze_configCondition(args,cfg):
    if args.attack!=None:

        list = args.attack.split(":")
        if len(list)==2:
            attack_type ,cfg.ATTACKMETHOD = list
        elif len(list)==3:
            attack_type ,cfg.ATTACKMETHOD = list[:2]
            cfg.ATTACKPRETRAINED = True
        else :
            raise ValueError(f'input :{args.attack} is not correct, please check usage in USE.md')

        if attack_type in attack_R_type:
            cfg.ATTACKTYPE = attack_type[:-1]
            cfg.ATTACKDIRECTION = attack_type[-1]
        else:
            raise ValueError(f"{attack_type} not found, please check usage in USE.md")
    if args.defense != None:
        list = args.defense.split(":")
        if len(list)==1:
            cfg.DEFENSEMETHOD = args.defense
        elif len(list)==2:
            cfg.DEFENSEMETHOD = list[0]
            cfg.DEFENSEPRETRAINED = True
        else:
            raise ValueError(f'too many :, please check usage in USE.md')
        cfg.MODEL.DEFENSE_TRAINED_WEIGHT = f'./model/{cfg.DEFENSEMETHOD}_{cfg.DATASETS.NAMES[0]}_{cfg.CFGTYPE}.pth'

    print('---------------------------------------------------------')
    print('---------------------------------------------------------')
    print('Option:')
    print(f"     attack : {args.attack!=None}")
    print(f"            attack  method    = {cfg.ATTACKMETHOD}")
    print(f"            attack  type      = {cfg.ATTACKTYPE}")
    print(f"            attack  direction = {cfg.ATTACKDIRECTION}")
    print(f"     defense: {args.defense!=None} ")
    print(f"            defense method    = {cfg.DEFENSEMETHOD}"  )
    print(f"     record : {args.record}")
    if args.record:
        print(f"            data will be recorded in ./{excel_name}")
    else:
        print(f"            data will not be recorded")

    print(f"     log    : {args.log}")
    if args.log:
        print(f"            data will be recorded in ./{excel_name}")
        if args.save_pic:
            print(f"            pictures will be saved")
        else:
            print(f"            pictures will not be saved")
    else:
        print(f"            data will not be recorded")
    print('---------------------------------------------------------')
    print('---------------------------------------------------------')
    return cfg