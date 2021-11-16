
import matplotlib.pyplot as plt
import openpyxl
import argparse
from openpyxl.utils import get_column_letter
import os
import time
# import sys
# sys.path.append('..')
# from fastreid.utils.reid_patch import Attack_algorithm_library,Defense_algorithm_library
excel_path = '../result.xlsx'
save_path = '../result'
C_Attack_algorithm_library=["FGSM",'IFGSM','MIFGSM','CW','ODFA']  #针对分类问题的攻击算法库
R_Attack_algorithm_library=['SMA','FNA','MIS-RANKING','MUAP','SSAE']
Attack_algorithm_library=C_Attack_algorithm_library+R_Attack_algorithm_library

G_Defense_algorithm_library=['ADV_DEF','GRA_REG','DISTILL']
R_Defense_algorithm_library=['GOAT']
Defense_algorithm_library=G_Defense_algorithm_library+R_Defense_algorithm_library

evaluation_indicator=['Rank-1','Rank-5','Rank-10','mAP','mINP','metric']

def main():
    args = default_argument_parser().parse_args()
    #print(args)
    wb = openpyxl.load_workbook(excel_path)
    sheet_list = wb.sheetnames
    assert (args.dataset in sheet_list),"please use the dataset in the excel, or check your spelling"
    sheet = wb[args.dataset]

    for i in args.attack:
        assert (i in Attack_algorithm_library),"please use the attack method in the excel, or check your spelling"
    for i in args.defense:
        assert (i in Defense_algorithm_library),"please use the defense method in the excel, or check your spelling"
    
    check_save_path(args)
    
    generate_images(args,sheet,args.attack,args.defense)


def check_save_path(args):
    global save_path
    if not os.path.isdir(save_path):
        os.mkdir(save_path)
    save_path = save_path+'/'+args.dataset
    if not os.path.isdir(save_path):
        os.mkdir(save_path)

def generate_images(args,sheet,attack,defense):
    plt.xlabel("stage")
    plt.ylabel(args.opt)
    plt.ylim([0,100])
    x=['pure','attack','defense','def-attack']
    bias = evaluation_indicator.index(args.opt)
    for attack_method in attack:
        y=[]
        index = Attack_algorithm_library.index(attack_method)
        row = 4 + len(evaluation_indicator)*index + bias
        y.append(round(sheet['D'+str(row)].value,1))
        y.append(round(sheet['E'+str(row)].value,1))
        for defense_method in defense:
            index = Defense_algorithm_library.index(defense_method)
            line_1 = get_column_letter(6+2*index)
            line_2 = get_column_letter(7+2*index)
            y.append(round(sheet[line_1+str(row)].value,1))
            y.append(round(sheet[line_2+str(row)].value,1))
            for a, b in zip(x, y):
                plt.text(a, b, b, fontsize=10)
            plt.plot(x,y,label=f"{attack_method}+{defense_method}")
            y.pop(-1)
            y.pop(-1)
    plt.legend()
    name = time.strftime("%H:%M:%S", time.localtime())
    name = name[0:2]+'_'+name[3:5]+'_'+name[6:8]
    plt.savefig(save_path+'/'+name +'.png')
    
    plt.show()
    plt.close()
    


def default_argument_parser():
    """
    Create a parser with some common arguments used by fastreid users.
    Returns:
        argparse.ArgumentParser:
    """
    parser = argparse.ArgumentParser(description="fastreid chart build")
    parser.add_argument("--dataset",default=None)
    parser.add_argument("--attack",default=None,nargs="+")
    parser.add_argument("--defense",default=None,nargs="+")
    parser.add_argument("--opt",default='Rank-1')
    return parser


















if __name__ =='__main__':
    main()